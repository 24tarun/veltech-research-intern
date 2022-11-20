#Ipothaiku ok
#Main~~~~~~!!!!!!!!!!

import numpy
import matplotlib.pyplot as plt
from pandas import read_csv
from pandas import read_excel
import pandas as pd
import math
from keras.models import Sequential
from keras.layers import Dense
from keras.layers import LSTM
from pandas.core.frame import DataFrame
from sklearn.preprocessing import MinMaxScaler
from sklearn.metrics import mean_squared_error
from openpyxl import Workbook, load_workbook
import gspread
from pprint import pprint



# convert an array of values into a dataset matrix
def create_dataset(dataset, look_back=1):
	dataX, dataY = [], []
	for i in range(len(dataset)-look_back-1):
		a = dataset[i:(i+look_back), 0]
		dataX.append(a)
		dataY.append(dataset[i + look_back, 0])
	return numpy.array(dataX), numpy.array(dataY)


# fix random seed for reproducibility
numpy.random.seed(7)

# load the dataset
#urlfile = "https://raw.githubusercontent.com/numenta/NAB/master/data/realKnownCause/ambient_temperature_system_failure.csv"

#dataframe = read_csv(r'C:\Users\philips1990\Desktop\Intern\Task 2\Anomaly_Detection\pH.csv', usecols=[1], engine='python')
#dataset = dataframe.values
#dataset = dataset.astype('float32')

#print(dataframe)
#print(dataset)



#excel_file = str(input("Enter the file name: "))
#dataframe = read_excel(r'C:\Users\philips1990\Desktop\Intern\Task 2\Anomaly_Detection\PH.xlsx')
#filedata = read_excel(excel_file)
#col_name = str(input("Enter the column name: "))
#pprint(filedata)

gc = gspread.service_account(filename = "creds.json")
#spreadsheet_id = "1gOCx6oGO_Tn5Wo27uUI2YH-3iNempLna8RRKnzOZ2xc"
spreadsheet_id = "1YQOSwpnr-b784AP8iVBZA_rJYB6HYla974Po8CtxthI"
sh = gc.open_by_key(spreadsheet_id).sheet1
#sh = gc.open_by_key(spreadsheet_id)
pprint(sh)
#list_of_lists = sh.get_all_values()
#pprint(list_of_lists)
#import numpy as np
array = numpy.array(sh.get_all_values())
#print(array[:10])
#print(array[3])
#print(sh.col_values(4))
#arr = array[0]

col_name = str(input("Enter the column name: "))

	
if col_name in array[0]:

        for ind in range(len(array[0])):
                if col_name == array[0][ind]:
                        col_val = ind+1
                        #print(col_val)
        
        dataframe = sh.col_values(col_val)
        dataframe = dataframe[1:]
        #dataset = dataframe.values
        dataset = numpy.array(dataframe).reshape(-1, 1)
        dataset = dataset.astype('float64')
        dataset = dataset[:1000]
        #print(dataframe)
        #print(dataset)
        #print(dataset[0])
        #print(dataset[0][0])
        #print(len(dataset))
        

        """Excel_name = input("Enter the file name for excel: ").strip()
        workbook = Workbook()
        sheet = workbook.active
        rowg = 1; colg = 1
        z = sheet.cell(row=rowg, column=colg)
        z.value = col_name
        rowg+=1
        for given_col_name in dataset:
                z = sheet.cell(row=rowg, column=colg)
                z.value = given_col_name[0]
                rowg += 1
        colg += 1
        workbook.save(filename=Excel_name+".xlsx")"""
        
        # normalize the dataset
        scaler = MinMaxScaler(feature_range=(0, 1))
        dataset = scaler.fit_transform(dataset)


        # split into train and test sets
        train_size = int(len(dataset) * 0.67)
        test_size = len(dataset) - train_size
        train, test = dataset[0:train_size,:], dataset[train_size:len(dataset),:]


        # reshape into X=t and Y=t+1
        look_back = 10
        trainX, trainY = create_dataset(train, look_back)
        testX, testY = create_dataset(test, look_back)


        # reshape input to be [samples, time steps, features]
        trainX = numpy.reshape(trainX, (trainX.shape[0], 1, trainX.shape[1]))
        testX = numpy.reshape(testX, (testX.shape[0], 1, testX.shape[1]))

        Epochs = int(input("Enter the number of epochs(25 recmm): "))
        # create and fit the LSTM network
        model = Sequential()
        model.add(LSTM(8, input_shape=(1, look_back)))
        model.add(Dense(1))
        model.compile(loss='mean_squared_error', optimizer='adam',metrics=['accuracy'])
        history = model.fit(trainX, trainY, epochs=Epochs, batch_size=32, verbose=2, validation_data=(trainX, trainY))
        print(history.history.keys())

        # make predictions
        trainPredict = model.predict(trainX)
        testPredict = model.predict(testX)

               
        # invert predictions
        trainPredict = scaler.inverse_transform(trainPredict)
        trainY = scaler.inverse_transform([trainY])
        testPredict = scaler.inverse_transform(testPredict)
        testY = scaler.inverse_transform([testY])


        # calculate root mean squared error
        trainScore = math.sqrt(mean_squared_error(trainY[0], trainPredict[:,0]))
        print('Train Score: %.2f RMSE' % (trainScore))
        testScore = math.sqrt(mean_squared_error(testY[0], testPredict[:,0]))
        print('Test Score: %.2f RMSE' % (testScore))


        # shift train predictions for plotting
        trainPredictPlot = numpy.empty_like(dataset)
        trainPredictPlot[:, :] = numpy.nan
        trainPredictPlot[look_back:len(trainPredict)+look_back, :] = trainPredict


        # shift test predictions for plotting
        testPredictPlot = numpy.empty_like(dataset)
        testPredictPlot[:, :] = numpy.nan
        testPredictPlot[len(trainPredict)+(look_back*2)+1:len(dataset)-1, :] = testPredict


        # plot baseline and predictions
        # trainPredictPlot
        plt.plot(scaler.inverse_transform(dataset), label = "Given Dataset")
        plt.plot(trainPredictPlot,label = "Training")
        plt.legend()
        plt.title('TrainPredictPlot')
        plt.ylabel(col_name)
        plt.xlabel('No of Values')
        plt.show()

        
        # testPredictPlot
        plt.plot(scaler.inverse_transform(dataset), label = "Given Dataset")
        plt.plot(testPredictPlot,label = "Predictions", color="red" )
        plt.legend()
        plt.title('TestPredictPlot')
        plt.ylabel(col_name)
        plt.xlabel('No of Values')
        plt.show()


        # Final Plot
        plt.plot(scaler.inverse_transform(dataset), label = "Given Dataset")
        plt.plot(trainPredictPlot,label = "Training")
        plt.plot(testPredictPlot, label = "Predictions")
        plt.legend()
        plt.title('Final Output')
        plt.ylabel(col_name)
        plt.xlabel('No of Values')
        plt.show()
        #print(trainPredictPlot)
        #print(testPredictPlot)

        # summarize history for accuracy
        plt.plot(history.history['accuracy'])
        plt.plot(history.history['val_accuracy'])
        plt.title('Model accuracy')
        plt.ylabel('Accuracy')
        plt.xlabel('Epoch')
        plt.legend(['Train Accuracy', 'Test Accuracy'], loc='upper right')
        plt.show()

        # summarize history for loss
        plt.plot(history.history['loss'])
        plt.plot(history.history['val_loss'])
        plt.title('Model Loss')
        plt.ylabel('Loss')
        plt.xlabel('Epoch')
        plt.legend(['Train Accuracy', 'Validation Accuracy'], loc='upper right')
        plt.show()


        """print("dataset",scaler.inverse_transform(dataset)[:10])
        print("Train",trainPredictPlot[:10])
        print("Test",testPredictPlot[:10])
        #print("Train",trainPredictPlot[0],trainPredictPlot[0][0])
        #print("Test",testPredictPlot[0],testPredictPlot[0][0])
        print("Accuracy",history.history['accuracy'][:10],len(history.history['accuracy']))
        print("Validation Accuracy",history.history['val_accuracy'][:10])
        print("Loss",history.history['loss'][:10])
        print("Validation Loss",history.history['val_loss'][:10])
        

        workbook = load_workbook(filename=Excel_name+".xlsx")
        sheet = workbook.active
        rowg = 1
        z = sheet.cell(row=rowg, column=colg)
        z.value = "TrainPredictVal"
        rowg+=1
        for train_val in trainPredictPlot:
                z = sheet.cell(row=rowg, column=colg)
                z.value = train_val[0]
                rowg += 1
        colg += 1
        rowg = 1
        z = sheet.cell(row=rowg, column=colg)
        z.value = "TestPredictVal"
        rowg+=1
        for test_val in testPredictPlot:
                z = sheet.cell(row=rowg, column=colg)
                z.value = test_val[0]
                rowg += 1
        colg += 1
        rowg = 1
        z = sheet.cell(row=rowg, column=colg)
        z.value = "Train Accuracy"
        rowg+=1
        for acc_val in history.history['accuracy']:
                z = sheet.cell(row=rowg, column=colg)
                z.value = acc_val#[0]
                rowg += 1
        colg += 1
        rowg = 1
        z = sheet.cell(row=rowg, column=colg)
        z.value = "Test Accuracy"
        rowg += 1
        for acc1_val in history.history['val_accuracy']:
                z = sheet.cell(row=rowg, column=colg)
                z.value = acc1_val#[0]
                rowg += 1
        colg += 1
        rowg = 1
        z = sheet.cell(row=rowg, column=colg)
        z.value = "Train Loss"
        rowg += 1
        for loss_val in history.history['loss']:
                z = sheet.cell(row=rowg, column=colg)
                z.value = loss_val#[0]
                rowg += 1
        colg += 1
        rowg = 1
        z = sheet.cell(row=rowg, column=colg)
        z.value = "Validation Loss"
        rowg += 1
        for val_loss in history.history['val_loss']:
                z = sheet.cell(row=rowg, column=colg)
                z.value = val_loss#[0]
                rowg += 1
        colg += 1
        rowg = 1
        z = sheet.cell(row=rowg, column=colg)
        z.value = "Epochs"
        rowg += 1
        for count in range(1,Epochs+1):
                z = sheet.cell(row=rowg, column=colg)
                z.value = count#[0]
                rowg += 1
        colg += 1
        workbook.save(filename=Excel_name+".xlsx")"""





else:
	print("Column not found -_-")

